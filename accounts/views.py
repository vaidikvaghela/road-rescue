from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Count, Sum
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import timedelta
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from .models import User
from .serializers import UserSerializer, RegisterSerializer, CustomTokenObtainPairSerializer


# ── HTML Pages ───────────────────────────────────────────────────────────────
def login_page(request):
    return render(request, 'accounts/login.html')

def register_page(request):
    return render(request, 'accounts/register.html')

def dashboard_page(request):
    return render(request, 'accounts/dashboard.html')

def admin_panel_page(request):
    return render(request, 'accounts/admin_panel.html')

def provider_onboarding_page(request):
    return render(request, 'accounts/provider_onboarding.html')


# ── REST API ─────────────────────────────────────────────────────────────────
class RegisterView(generics.CreateAPIView):
    queryset           = User.objects.all()
    serializer_class   = RegisterSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        # Generate tokens immediately so providers can go straight to onboarding
        from rest_framework_simplejwt.tokens import RefreshToken
        refresh = RefreshToken.for_user(user)
        return Response({
            'message': 'Registration successful.',
            'user':    UserSerializer(user).data,
            'tokens':  {
                'access':  str(refresh.access_token),
                'refresh': str(refresh),
            }
        }, status=status.HTTP_201_CREATED)

class LoginView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

class ProfileView(generics.RetrieveUpdateAPIView):
    serializer_class   = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user

class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user         = request.user
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        if not user.check_password(old_password):
            return Response({'error': 'Incorrect current password.'}, status=400)
        user.set_password(new_password)
        user.save()
        return Response({'message': 'Password updated successfully.'})


class ProviderOnboardingView(APIView):
    """Lets a newly registered provider complete their ServiceProvider profile."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Return current provider profile (or stub info)."""
        if request.user.role != 'provider':
            return Response({'error': 'Not a provider account.'}, status=403)
        from services.models import ServiceProvider, ServiceType
        from services.serializers import ServiceProviderSerializer, ServiceTypeSerializer
        try:
            profile = request.user.provider_profile
            data = ServiceProviderSerializer(profile, context={'request': request}).data
        except Exception:
            data = {}
        types = ServiceType.objects.filter(is_active=True)
        return Response({
            'profile': data,
            'service_types': ServiceTypeSerializer(types, many=True).data,
        })

    def post(self, request):
        """Create or update the ServiceProvider profile."""
        if request.user.role != 'provider':
            return Response({'error': 'Not a provider account.'}, status=403)
        from services.models import ServiceProvider, ServiceType
        from services.serializers import ServiceProviderSerializer

        profile, _ = ServiceProvider.objects.get_or_create(
            user=request.user,
            defaults={'phone': '', 'email': request.user.email,
                      'address': '', 'city': '', 'state': '', 'pincode': '',
                      'business_name': request.user.username + "'s Service"}
        )

        fields = ['business_name', 'description', 'phone', 'email',
                  'address', 'city', 'state', 'pincode',
                  'latitude', 'longitude', 'is_24_7', 'base_fee',
                  'opening_time', 'closing_time', 'is_available']
        for f in fields:
            if f in request.data:
                setattr(profile, f, request.data[f])

        # Handle service_types (list of IDs)
        service_type_ids = request.data.get('service_types', [])
        if service_type_ids:
            profile.service_types.set(ServiceType.objects.filter(id__in=service_type_ids))

        profile.status = 'active'
        profile.save()

        return Response({
            'message': 'Profile updated successfully.',
            'profile': ServiceProviderSerializer(profile, context={'request': request}).data,
        })


# ── Admin API ─────────────────────────────────────────────────────────────────
class AdminStatsView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        from emergency.models import EmergencyRequest, Payment
        from services.models import ServiceProvider

        today = timezone.now().date()
        week_ago = today - timedelta(days=6)

        # ── Top-level counts ──────────────────────────────────────────────────
        total_users     = User.objects.count()
        total_requests  = EmergencyRequest.objects.count()
        total_providers = ServiceProvider.objects.count()
        total_revenue   = Payment.objects.filter(status='completed').aggregate(total=Sum('amount'))['total'] or 0

        # ── Requests by status (donut chart) ─────────────────────────────────
        req_by_status = list(
            EmergencyRequest.objects.values('status').annotate(count=Count('id'))
        )

        # ── Revenue last 7 days (line chart) ─────────────────────────────────
        revenue_qs = (
            Payment.objects
            .filter(status='completed', created_at__date__gte=week_ago)
            .annotate(day=TruncDate('created_at'))
            .values('day')
            .annotate(total=Sum('amount'))
            .order_by('day')
        )
        revenue_by_day = {str(r['day']): float(r['total']) for r in revenue_qs}
        revenue_labels, revenue_data = [], []
        for i in range(7):
            d = str(week_ago + timedelta(days=i))
            revenue_labels.append(d)
            revenue_data.append(revenue_by_day.get(d, 0))

        # ── User signups last 7 days (bar chart) ─────────────────────────────
        signup_qs = (
            User.objects
            .filter(date_joined__date__gte=week_ago)
            .annotate(day=TruncDate('date_joined'))
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )
        signup_by_day = {str(r['day']): r['count'] for r in signup_qs}
        signup_labels, signup_data = [], []
        for i in range(7):
            d = str(week_ago + timedelta(days=i))
            signup_labels.append(d)
            signup_data.append(signup_by_day.get(d, 0))

        # ── Issue type breakdown (horizontal bar) ────────────────────────────
        issue_qs = (
            EmergencyRequest.objects.values('issue_type').annotate(count=Count('id')).order_by('-count')
        )
        issue_labels = [r['issue_type'].replace('_', ' ').title() for r in issue_qs]
        issue_data   = [r['count'] for r in issue_qs]

        # ── Recent requests table ─────────────────────────────────────────────
        recent_requests = list(
            EmergencyRequest.objects.select_related('assigned_provider').values(
                'id', 'requester_name', 'requester_phone', 'issue_type', 'vehicle_type',
                'status', 'priority', 'location_address', 'created_at'
            ).order_by('-created_at')[:20]
        )
        for r in recent_requests:
            r['created_at'] = r['created_at'].strftime('%d %b %Y, %H:%M')

        # ── Recent users table ────────────────────────────────────────────────
        recent_users = list(
            User.objects.values('id', 'email', 'username', 'first_name', 'last_name', 'role', 'is_staff', 'date_joined').order_by('-date_joined')[:20]
        )
        for u in recent_users:
            u['date_joined'] = u['date_joined'].strftime('%d %b %Y')

        # ── Recent payments table ─────────────────────────────────────────────
        recent_payments = list(
            Payment.objects.select_related('request').values(
                'id', 'amount', 'status', 'razorpay_order_id', 'razorpay_payment_id', 'created_at',
                'request__id', 'request__requester_name'
            ).order_by('-created_at')[:20]
        )
        for p in recent_payments:
            p['created_at'] = p['created_at'].strftime('%d %b %Y, %H:%M')
            p['amount'] = float(p['amount'])

        # ── Recent providers table ────────────────────────────────────────────
        recent_providers = list(
            ServiceProvider.objects.values(
                'id', 'business_name', 'city', 'status', 'rating', 'total_jobs', 'is_available', 'created_at'
            ).order_by('-created_at')[:20]
        )
        for p in recent_providers:
            p['created_at'] = p['created_at'].strftime('%d %b %Y')
            p['rating'] = float(p['rating'])

        return Response({
            'summary': {
                'total_users':     total_users,
                'total_requests':  total_requests,
                'total_providers': total_providers,
                'total_revenue':   float(total_revenue),
            },
            'charts': {
                'req_by_status':  req_by_status,
                'revenue_labels': revenue_labels,
                'revenue_data':   revenue_data,
                'commission_data': [val * 0.20 for val in revenue_data],
                'signup_labels':  signup_labels,
                'signup_data':    signup_data,
                'issue_labels':   issue_labels,
                'issue_data':     issue_data,
            },
            'tables': {
                'requests':  recent_requests,
                'users':     recent_users,
                'payments':  recent_payments,
                'providers': recent_providers,
            }
        })


class AdminExportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        from emergency.models import EmergencyRequest, Payment
        from services.models import ServiceProvider

        model = request.query_params.get('model', 'requests')

        wb = openpyxl.Workbook()
        ws = wb.active

        # Styling helpers
        header_font    = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
        header_fill    = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        center_align   = Alignment(horizontal='center', vertical='center')
        thin_border    = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        alt_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')

        def style_header(ws, headers):
            ws.append(headers)
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border

        def style_row(ws, row_idx):
            fill = alt_fill if row_idx % 2 == 0 else None
            for cell in ws[row_idx]:
                cell.alignment = center_align
                cell.border    = thin_border
                if fill:
                    cell.fill = fill

        if model == 'requests':
            ws.title = 'Emergency Requests'
            style_header(ws, ['ID', 'Requester', 'Phone', 'Issue', 'Vehicle', 'Status', 'Priority', 'Location', 'Created At'])
            for i, r in enumerate(EmergencyRequest.objects.all().order_by('-created_at'), start=2):
                ws.append([
                    r.pk, r.requester_name, r.requester_phone,
                    r.get_issue_type_display(), r.get_vehicle_type_display(),
                    r.get_status_display(), r.get_priority_display(),
                    r.location_address, r.created_at.strftime('%d %b %Y %H:%M')
                ])
                style_row(ws, i)
            filename = 'emergency_requests.xlsx'

        elif model == 'users':
            ws.title = 'Users'
            style_header(ws, ['ID', 'Email', 'Username', 'First Name', 'Last Name', 'Role', 'Staff', 'Verified', 'Joined'])
            for i, u in enumerate(User.objects.all().order_by('-date_joined'), start=2):
                ws.append([
                    u.pk, u.email, u.username, u.first_name, u.last_name,
                    u.role, 'Yes' if u.is_staff else 'No',
                    'Yes' if u.is_verified else 'No',
                    u.date_joined.strftime('%d %b %Y')
                ])
                style_row(ws, i)
            filename = 'users.xlsx'

        elif model == 'payments':
            ws.title = 'Payments'
            style_header(ws, ['ID', 'Request ID', 'Requester', 'Amount (INR)', 'Status', 'Razorpay Order ID', 'Razorpay Payment ID', 'Created At'])
            for i, p in enumerate(Payment.objects.select_related('request').all().order_by('-created_at'), start=2):
                ws.append([
                    p.pk, p.request.pk, p.request.requester_name,
                    float(p.amount), p.get_status_display(),
                    p.razorpay_order_id or '—', p.razorpay_payment_id or '—',
                    p.created_at.strftime('%d %b %Y %H:%M')
                ])
                style_row(ws, i)
            filename = 'payments.xlsx'

        elif model == 'providers':
            ws.title = 'Service Providers'
            style_header(ws, ['ID', 'Business Name', 'City', 'State', 'Status', 'Rating', 'Total Jobs', 'Available', 'Created At'])
            from services.models import ServiceProvider
            for i, p in enumerate(ServiceProvider.objects.all().order_by('-created_at'), start=2):
                ws.append([
                    p.pk, p.business_name, p.city, p.state,
                    p.get_status_display(), float(p.rating), p.total_jobs,
                    'Yes' if p.is_available else 'No',
                    p.created_at.strftime('%d %b %Y')
                ])
                style_row(ws, i)
            filename = 'service_providers.xlsx'

        else:
            return Response({'error': 'Invalid model parameter.'}, status=400)

        # Auto-fit columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class AdminActionView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request):
        action = request.data.get('action')
        model = request.data.get('model')
        item_id = request.data.get('id')

        try:
            if model == 'user':
                user = User.objects.get(id=item_id)
                if action == 'delete':
                    if user.is_superuser:
                        return Response({'error': 'Cannot delete superuser'}, status=400)
                    user.delete()
                    return Response({'message': 'User deleted'})
                elif action == 'toggle_staff':
                    user.is_staff = not user.is_staff
                    user.save(update_fields=['is_staff'])
                    return Response({'message': f'Staff status set to {user.is_staff}'})

            elif model == 'provider':
                from services.models import ServiceProvider
                provider = ServiceProvider.objects.get(id=item_id)
                if action == 'approve':
                    provider.status = 'active'
                    provider.save(update_fields=['status'])
                    return Response({'message': 'Provider approved'})
                elif action == 'suspend':
                    provider.status = 'suspended'
                    provider.save(update_fields=['status'])
                    return Response({'message': 'Provider suspended'})
                elif action == 'delete':
                    provider.delete()
                    return Response({'message': 'Provider deleted'})

            elif model == 'request':
                from emergency.models import EmergencyRequest
                req = EmergencyRequest.objects.get(id=item_id)
                if action == 'cancel':
                    req.status = 'cancelled'
                    req.save(update_fields=['status'])
                    return Response({'message': 'Request cancelled'})
                elif action == 'delete':
                    req.delete()
                    return Response({'message': 'Request deleted'})

            elif model == 'payment':
                from emergency.models import Payment
                payment = Payment.objects.get(id=item_id)
                if action == 'refund':
                    payment.status = 'refunded'
                    payment.save(update_fields=['status'])
                    return Response({'message': 'Payment refunded'})

            return Response({'error': 'Invalid action or model'}, status=400)
        except Exception as e:
            return Response({'error': str(e)}, status=400)
